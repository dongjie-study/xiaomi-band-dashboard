import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from collections import defaultdict

# ============================================================
# DATA PREPARATION
# ============================================================

with open(r'C:\Users\Administrator\Desktop\anchor_summary.json', 'r', encoding='utf-8') as f:
    anchor_data = json.load(f)

with open(r'C:\Users\Administrator\Desktop\小米手环直播间销量分析\sales_analysis\history.json', 'r', encoding='utf-8') as f:
    history = json.load(f)

july_data = [d for d in history if d.get('date','').startswith('2026-07')]

# ============================================================
# WORKBOOK & STYLES
# ============================================================

wb = Workbook()

title_font = Font(name='Arial', size=16, bold=True, color='1a1a2e')
header_font = Font(name='Arial', size=11, bold=True, color='ffffff')
normal_font = Font(name='Arial', size=10, color='333333')
bold_font = Font(name='Arial', size=10, bold=True, color='333333')
highlight_font = Font(name='Arial', size=10, bold=True, color='ff6900')
big_num_font = Font(name='Arial', size=12, bold=True, color='ff6900')
small_font = Font(name='Arial', size=9, color='666666')
white_font = Font(name='Arial', size=10, color='ffffff')

title_fill = PatternFill('solid', fgColor='1a1a2e')
header_fill = PatternFill('solid', fgColor='ff6900')
header_fill2 = PatternFill('solid', fgColor='0f3460')
alt_fill = PatternFill('solid', fgColor='f8fafc')
gold_fill = PatternFill('solid', fgColor='fff7ed')
green_fill = PatternFill('solid', fgColor='f0faf3')
red_fill = PatternFill('solid', fgColor='fef2f2')
blue_fill = PatternFill('solid', fgColor='eff6ff')
light_gray_fill = PatternFill('solid', fgColor='f5f5f5')
white_fill = PatternFill('solid', fgColor='ffffff')

thin_border = Border(
    left=Side(style='thin', color='e0e0e0'),
    right=Side(style='thin', color='e0e0e0'),
    top=Side(style='thin', color='e0e0e0'),
    bottom=Side(style='thin', color='e0e0e0')
)

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

def style_header_row(ws, row, max_col, fill=header_fill, font=header_font):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font
        cell.fill = fill
        cell.alignment = center_align
        cell.border = thin_border

def sc(ws, row, col, value, font=normal_font, align=center_align, nf=None, fill=None):
    """Style a data cell"""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = font
    cell.alignment = align
    cell.border = thin_border
    if nf:
        cell.number_format = nf
    if fill:
        cell.fill = fill
    return cell

def add_title(ws, row, title, max_col):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).fill = title_fill
        ws.cell(row=row, column=c).border = thin_border
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 40
    return row + 1

def add_subtitle(ws, row, subtitle, max_col):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).fill = light_gray_fill
        ws.cell(row=row, column=c).border = thin_border
    cell = ws.cell(row=row, column=1, value=subtitle)
    cell.font = Font(name='Arial', size=11, color='666666')
    cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row].height = 25
    return row + 1

def add_section_header(ws, row, title, max_col, color='0f3460'):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=max_col)
    for c in range(1, max_col + 1):
        ws.cell(row=row, column=c).fill = PatternFill('solid', fgColor=color)
        ws.cell(row=row, column=c).border = thin_border
    cell = ws.cell(row=row, column=1, value=title)
    cell.font = Font(name='Arial', size=13, bold=True, color='ffffff')
    cell.alignment = left_align
    ws.row_dimensions[row].height = 32
    return row + 1

# ============================================================
# SHEET 1: Anchor Performance Summary
# ============================================================
ws1 = wb.active
ws1.title = '主播业绩汇总'
MAX_C1 = 11

row = 1
row = add_title(ws1, row, '小米官方耳机直播间(我司) 7月主播业绩复盘', MAX_C1)
row = add_subtitle(ws1, row, f'数据周期: 2026.7.1-7.18 | 全直播间总GSV: {anchor_data["total_gsv"]:,.0f} | 主播: 9人 | 总开播: 49人次', MAX_C1)

# KPI banner
row += 1
kpi_items = [
    ('全直播间总GSV', f'{anchor_data["total_gsv"]:,.0f}'),
    ('主播总人数', '9人'),
    ('总开播人次', '49人次'),
    ('场均GSV', f'{anchor_data["total_gsv"]/49:,.0f}'),
]
for i, (label, value) in enumerate(kpi_items):
    c1 = i * 3 + 1
    c2 = c1 + 1
    ws1.merge_cells(start_row=row, start_column=c1, end_row=row, end_column=c2)
    # Write before merge to avoid MergedCell issue
    sc(ws1, row, c1, label, font=Font(name='Arial', size=9, bold=True, color='ff6900'), fill=gold_fill)
    # After merge, the second cell becomes MergedCell. Just style it.
    for cc in range(c1, c2+1):
        ws1.cell(row=row, column=cc).fill = gold_fill
        ws1.cell(row=row, column=cc).border = thin_border
    # Write value in the first cell (overwrites label)
    sc(ws1, row, c1, f'{label}\n{value}', font=Font(name='Arial', size=14, bold=True, color='ff6900'), fill=gold_fill)
ws1.row_dimensions[row].height = 45

row += 2

# Main table
headers = ['排名', '主播', '总GSV', '开播天数', '有业绩天数', '日均GSV(有业绩)', '单日最高GSV', '最高日期', '占比', '评级', '核心亮点/问题']
for i, h in enumerate(headers):
    sc(ws1, row, i+1, h, font=header_font, fill=header_fill)
ws1.row_dimensions[row].height = 35
row += 1

sorted_anchors = sorted(anchor_data['anchor_summary'], key=lambda x: x['总GSV'], reverse=True)
for rank, a in enumerate(sorted_anchors):
    total = a['总GSV']
    avg = a['日均GSV']
    max_gsv = a['最高GSV']

    if total >= 50000:
        rating, rating_fill, rating_color = 'S', gold_fill, 'ff6900'
    elif total >= 10000:
        rating, rating_fill, rating_color = 'A', green_fill, '1da85c'
    elif total >= 3000:
        rating, rating_fill, rating_color = 'B', blue_fill, '3b82f6'
    else:
        rating, rating_fill, rating_color = 'C', light_gray_fill, '999999'

    # Key insight for each anchor
    insights_map = {
        '徐梦洁': '日均26K, 7.17单人爆发89K, 转化能力顶级, 需稳定排班, 大促必排',
        '刘子源': '日均18.6K, 7.17单场60K, 大场型主播, 适合流量高峰期',
        '王瑞': '最勤主播(12天), 稳定输出日均3.8K, 平播中坚力量',
        '李晓洋': '12天出勤, 7.18达10K, 逐步提升中, 需加强逼单话术',
        '范金蝶': '7.9单场15K亮眼, 但后期排班减少, 建议增加场次',
        '陈依雨': '前期稳定, 7.7后无排班, 需确认是否流失或转岗',
        '于飞': '仅1天试播2.6K, 需更多排班验证能力',
        '郑涵悦': '仅1天试播1.7K, 新人需师父带教',
        '王洋帆': '仅1天试播861, 新人起步阶段'
    }
    highlight_text = insights_map.get(a['主播'], '')

    row_fill = alt_fill if rank % 2 == 0 else white_fill
    if rank < 2:
        row_fill = gold_fill

    vals = [rank+1, a['主播'], total, a['开播天数'], a['有业绩天数'], avg, max_gsv, a['最高日期'], f"{a['占比']}%", rating, highlight_text]
    for i, v in enumerate(vals):
        if i == 0:
            sc(ws1, row, i+1, v, font=bold_font, fill=row_fill)
        elif i == 2:
            sc(ws1, row, i+1, v, font=big_num_font, nf='#,##0', fill=row_fill)
        elif i in (5, 6):
            sc(ws1, row, i+1, v, nf='#,##0', fill=row_fill)
        elif i == 9:
            sc(ws1, row, i+1, v, font=Font(name='Arial', size=11, bold=True, color=rating_color), fill=rating_fill)
        elif i == 10:
            sc(ws1, row, i+1, v, font=small_font, align=left_align, fill=row_fill)
        else:
            sc(ws1, row, i+1, v, fill=row_fill)
    ws1.row_dimensions[row].height = 30
    row += 1

# Totals row
total_days = sum(a['开播天数'] for a in sorted_anchors)
total_active = sum(a['有业绩天数'] for a in sorted_anchors)
sc(ws1, row, 1, '', fill=light_gray_fill)
sc(ws1, row, 2, '合计', font=Font(name='Arial', size=12, bold=True), fill=light_gray_fill)
sc(ws1, row, 3, anchor_data['total_gsv'], font=Font(name='Arial', size=13, bold=True, color='ff6900'), nf='#,##0', fill=light_gray_fill)
sc(ws1, row, 4, total_days, font=bold_font, fill=light_gray_fill)
sc(ws1, row, 5, total_active, font=bold_font, fill=light_gray_fill)
sc(ws1, row, 6, anchor_data['total_gsv']/total_active if total_active else 0, font=bold_font, nf='#,##0', fill=light_gray_fill)
sc(ws1, row, 7, max(a['最高GSV'] for a in sorted_anchors), font=highlight_font, nf='#,##0', fill=light_gray_fill)
for i in range(8, MAX_C1+1):
    sc(ws1, row, i, '', fill=light_gray_fill)

# Column widths
widths1 = [6, 10, 14, 10, 10, 14, 14, 10, 8, 8, 50]
for i, w in enumerate(widths1):
    ws1.column_dimensions[get_column_letter(i+1)].width = w

# Key findings
row += 2
row = add_section_header(ws1, row, '关键发现与建议', MAX_C1)

findings = [
    '1. 主播梯队分明: 徐梦洁(104K)和刘子源(93K)贡献63%的GSV, 是绝对核心。两人适合大促/爆品期排班, 流量高峰期应优先安排。',
    '2. 爆发日集中: 7.17(160K)和7.18(51K)贡献67%的月度GSV, Xiaomi Buds 6爆品+大主播组合效果显著, 应复制此模式。',
    '3. 腰部主播断层: 王瑞和李晓洋各开播12天, 但日均仅3.2K-3.8K, 与头部差距8倍, 急需话术培训+逼单能力提升。',
    '4. 新人待培养: 于飞/郑涵悦/王洋帆仅各播1天, 需安排师父带教+旁听3天后再上岗, 逐步增加时长。',
    '5. 排班优化: 7.13-7.15无数据(可能未开播), 范金蝶7.9爆15K后消失, 陈依雨7.7后消失, 排班合理性需审视。',
    '6. 与良米差距: 我司仅占耳机直播间总GSV的17%, 良米以6倍订单量领先。但我司客单价474 vs 良米380, 高端产品线是优势方向。',
]
for f in findings:
    ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_C1)
    sc(ws1, row, 1, f, font=Font(name='Arial', size=10), align=left_align)
    ws1.row_dimensions[row].height = 25
    row += 1

row += 1
ws1.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_C1)
sc(ws1, row, 1, '注意: 7.13-7.15数据缺失, 需从百应后台补录。主播GSV截至7.18, 7.19数据待补充。', font=Font(name='Arial', size=9, color='999999', italic=True), align=left_align)

print('Sheet 1 done')

# ============================================================
# SHEET 2: Daily Detail
# ============================================================
ws2 = wb.create_sheet('每日业绩明细')
MAX_C2 = 9

row = 1
row = add_title(ws2, row, '每日业绩明细 & 趋势分析', MAX_C2)
row = add_subtitle(ws2, row, '7月1日-7月18日逐日GSV趋势 | 含环比变化、主力主播、趋势判断', MAX_C2)

headers2 = ['日期', '总GSV', '主播数', '出单主播数', '场均GSV', '环比变化', '主力主播(Top3)', '趋势', '备注']
for i, h in enumerate(headers2):
    sc(ws2, row, i+1, h, font=header_font, fill=header_fill2)
ws2.row_dimensions[row].height = 30
row += 1

daily_data = defaultdict(lambda: {'gsv': 0, 'anchors': set(), 'active': []})
for r in anchor_data['records']:
    d = r['日期']
    daily_data[d]['gsv'] += r['GSV']
    daily_data[d]['anchors'].add(r['主播'])
    if r['GSV'] > 0:
        daily_data[d]['active'].append((r['主播'], r['GSV']))

sorted_dates = sorted(daily_data.keys(), key=lambda x: float(x))
prev_gsv = None

for date_str in sorted_dates:
    d = daily_data[date_str]
    gsv = d['gsv']
    n_anchors = len(d['anchors'])
    n_active = len(d['active'])
    avg_gsv = gsv / n_active if n_active > 0 else 0

    if prev_gsv and prev_gsv > 0:
        change = (gsv - prev_gsv) / prev_gsv * 100
        change_str = f'+{change:.0f}%' if change >= 0 else f'{change:.0f}%'
    else:
        change_str = '-'
    prev_gsv = gsv if gsv > 0 else prev_gsv

    d['active'].sort(key=lambda x: x[1], reverse=True)
    top_anchors = ', '.join([f'{a}({g:,.0f})' for a, g in d['active'][:3]]) if d['active'] else '-'

    if gsv == 0:
        trend, trend_fill_val = '未开播/零业绩', 'fef2f2'
    elif gsv >= 50000:
        trend, trend_fill_val = '爆发日', 'fff7ed'
    elif gsv >= 10000:
        trend, trend_fill_val = '良好', 'f0faf3'
    elif gsv >= 3000:
        trend, trend_fill_val = '平播', 'eff6ff'
    else:
        trend, trend_fill_val = '低迷', 'fef2f2'

    df_val = float(date_str)
    if df_val >= 7.17:
        note = '大促/爆品期 - Buds 6起量'
    elif df_val >= 7.8:
        note = '中期稳定增长'
    elif df_val >= 7.3:
        note = '起步阶段'
    else:
        note = '冷启动期'

    rf = alt_fill if sorted_dates.index(date_str) % 2 == 0 else white_fill
    vals = [f'7月{date_str.split(".")[1]}日', gsv, n_anchors, n_active, avg_gsv, change_str, top_anchors, trend, note]
    for i, v in enumerate(vals):
        if i == 1:
            sc(ws2, row, i+1, v, font=big_num_font, nf='#,##0', fill=rf)
        elif i == 4:
            sc(ws2, row, i+1, v, nf='#,##0', fill=rf)
        elif i == 6:
            sc(ws2, row, i+1, v, font=small_font, align=left_align, fill=rf)
        elif i == 7:
            sc(ws2, row, i+1, v, font=Font(name='Arial', size=9, bold=True), fill=PatternFill('solid', fgColor=trend_fill_val))
        elif i == 8:
            sc(ws2, row, i+1, v, font=small_font, align=left_align, fill=rf)
        else:
            sc(ws2, row, i+1, v, fill=rf)
    ws2.row_dimensions[row].height = 26
    row += 1

widths2 = [10, 14, 8, 10, 14, 10, 40, 14, 28]
for i, w in enumerate(widths2):
    ws2.column_dimensions[get_column_letter(i+1)].width = w

# Phase summary
row += 1
row = add_section_header(ws2, row, '月度阶段分析', MAX_C2)

phases = [
    ('阶段一 (7.1-7.7) 冷启动起步期', '日均GSV 8,071 | 王瑞为主力, 李晓洋/范金蝶/陈依雨辅助, 逐步建立直播节奏'),
    ('阶段二 (7.8-7.12) 调整期', '日均GSV 6,828 | 范金蝶7.9爆发15K, 王瑞持续稳定, 新人郑涵悦/王洋帆首次试播'),
    ('阶段三 (7.13-7.15) 数据缺失', '无数据 - 需从百应后台确认, 如未开播则为排班空窗期'),
    ('阶段四 (7.16-7.18) 爆发期', '日均GSV 73,266 | 徐梦洁+刘子源双引擎, Xiaomi Buds 6爆品驱动, 7.17单日160K创纪录'),
]
for title, desc in phases:
    ws2.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_C2)
    sc(ws2, row, 1, f'{title}: {desc}', font=Font(name='Arial', size=10), align=left_align)
    ws2.row_dimensions[row].height = 22
    row += 1

print('Sheet 2 done')

# ============================================================
# SHEET 3: GSV Matrix
# ============================================================
ws3 = wb.create_sheet('GSV矩阵')

pivot_df = pd.DataFrame(anchor_data['records'])
pivot = pivot_df.pivot_table(values='GSV', index='日期', columns='主播', aggfunc='sum', fill_value=0)
pivot = pivot.sort_index(key=lambda x: [float(i) for i in x])
col_totals = pivot.sum()
sorted_cols = col_totals.sort_values(ascending=False).index.tolist()
pivot = pivot[sorted_cols]
MAX_C3 = len(sorted_cols) + 2

row = 1
row = add_title(ws3, row, '主播 x 日期 GSV矩阵', MAX_C3)
row += 1

sc(ws3, row, 1, '日期', font=header_font, fill=header_fill)
for i, col in enumerate(sorted_cols):
    sc(ws3, row, i+2, col, font=header_font, fill=header_fill)
sc(ws3, row, MAX_C3, '日合计', font=header_font, fill=header_fill)
ws3.row_dimensions[row].height = 30
row += 1

for date in pivot.index:
    rf = alt_fill if pivot.index.get_loc(date) % 2 == 0 else white_fill
    sc(ws3, row, 1, f'7月{date.split(".")[1]}日', fill=rf)
    row_total = 0
    for i, col in enumerate(sorted_cols):
        val = pivot.loc[date, col]
        row_total += val
        if val > 0:
            if val >= 10000:
                cf, cfont = gold_fill, Font(name='Arial', size=10, bold=True, color='ff6900')
            elif val >= 3000:
                cf, cfont = green_fill, Font(name='Arial', size=10, bold=True, color='1da85c')
            else:
                cf, cfont = rf, normal_font
            sc(ws3, row, i+2, val, font=cfont, nf='#,##0', fill=cf)
        else:
            sc(ws3, row, i+2, '-', font=Font(name='Arial', size=9, color='cccccc'), fill=rf)

    total_fill_val = gold_fill if row_total >= 50000 else (green_fill if row_total >= 10000 else rf)
    sc(ws3, row, MAX_C3, row_total, font=big_num_font, nf='#,##0', fill=total_fill_val)
    ws3.row_dimensions[row].height = 26
    row += 1

# Column totals
sc(ws3, row, 1, '合计', font=bold_font, fill=light_gray_fill)
grand = 0
for i, col in enumerate(sorted_cols):
    ct = pivot[col].sum()
    grand += ct
    sc(ws3, row, i+2, ct, font=highlight_font, nf='#,##0', fill=light_gray_fill)
sc(ws3, row, MAX_C3, grand, font=Font(name='Arial', size=13, bold=True, color='ff6900'), nf='#,##0', fill=light_gray_fill)

ws3.column_dimensions['A'].width = 12
for i in range(len(sorted_cols)):
    ws3.column_dimensions[get_column_letter(i+2)].width = 12
ws3.column_dimensions[get_column_letter(MAX_C3)].width = 14

row += 2
ws3.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_C3)
sc(ws3, row, 1, '颜色: 橙色=10K+ | 绿色=3-10K | 白色=1-3K | 灰色=-', font=small_font, align=left_align)

print('Sheet 3 done')

# ============================================================
# SHEET 4: 话术SOP
# ============================================================
ws4 = wb.create_sheet('话术SOP')
MAX_C4 = 4

row = 1
row = add_title(ws4, row, '耳机直播间 话术SOP标准流程', MAX_C4)
row = add_subtitle(ws4, row, '基于高转化主播(徐梦洁/刘子源)实战经验提炼 | 7步标准话术框架', MAX_C4)

# SOP Table
row += 1
sop_headers = ['步骤', '话术模块', '标准话术要点', '时间分配']
for i, h in enumerate(sop_headers):
    sc(ws4, row, i+1, h, font=header_font, fill=header_fill2)
ws4.row_dimensions[row].height = 30
row += 1

sop_items = [
    ('Step 1', '开播暖场 + 适配度打消顾虑',
     '"直播间有多少小米手机/非小米手机的宝宝？不管OPPO/vivo/华为/苹果, 只要是智能手机有蓝牙, 咱们耳机都能正常连接使用。小米怎么连苹果安卓就怎么连, 大家放心拍。"\n\n【关键点】在所有产品讲解前, 先回答最高频疑问, 消除购买门槛。',
     '1-2分钟'),
    ('Step 2', '产品功能快速塑品 (快速转化版)',
     '"Xiaomi Buds 6: 半入耳主动降噪, 独立空间音频+头部追踪。Hi-Res认证+LDHC 5.0编解码, 音质对标千元级。单耳续航10小时, 总续航40小时。双设备无缝切换, 支持查找耳机、IP54防水。小黄车X号链接, 现在拍今天发顺丰。"\n\n【关键点】30秒内讲完核心卖点, 适合已经了解产品、比价的用户快速成交。',
     '30秒-1分钟'),
    ('Step 3', '深度塑品 (新客转化版)',
     '"来宝宝们, 我给你们详细讲一下这款耳机为什么值得买:\n\n①音质: Hi-Res金标认证+LDHC 5.0, 比普通蓝牙耳机传输信息量多3倍, 高音通透低音饱满\n②降噪: 半入耳也能主动降噪, 地铁公交上自动过滤噪音, 办公室专注模式下隔音效果超好\n③空间音频: 头部追踪+独立空间音频, 看电影打游戏身临其境, 千元以下唯一支持独立空间音频的耳机\n④续航: 单耳10小时(降噪关)/6小时(降噪开), 充电盒再续30小时, 一周一充\n⑤细节: 双设备无缝切换(手机+电脑), IP54防汗防水, 查找耳机, Google Fast Pair弹窗\n\n现在活动价只要XXX, 比其他平台便宜XX, 而且是官方直播间正品保障。"\n\n【关键点】按音质/降噪/空间音频/续航/细节5个维度分层讲解, 每个功能用场景化描述。',
     '3-5分钟'),
    ('Step 4', '产品对比引导 (推高客单价)',
     '"如果预算有限选REDMI Buds 8青春版(XXX), 追求性价比选REDMI Buds 8 Pro(XXX), 想要旗舰体验一步到位选Xiaomi Buds 6。"\n\n对比话术:\n- Buds 8青春版: 基础款, 适合听歌接电话够用就行\n- Buds 8 Pro: 加主动降噪+LDHC, 音质和降噪提升明显, 性价比最高\n- Buds 6: 半入耳旗舰, 空间音频+头部追踪, 降噪/音质/体验都是顶配\n\n"【关键点】从低到高3个价位段对比, 用\'加XX元多XX功能\'的逻辑推高客单价。',
     '1-2分钟'),
    ('Step 5', '互动确认 + 售后锁单',
     '"要拍的宝宝打\'要\'让我看到！拍了的小宝全屏飘\'已拍\', 主播给你们优先发货。\n\n我们是小米官方直播间:\n①7天无理由, 激活后不满意也能退\n②1年质保全国联保, 任何质量问题去就近小米门店免费修\n③发顺丰/京东, 今天拍明天到\n④正品保障, 官方发票\n\n【关键点】4重保障降低决策门槛, 用官方身份建立信任。',
     '1分钟'),
    ('Step 6', '库存/优惠逼单',
     '"现在库存不多了, XX颜色只剩最后XX个, 这批卖完下一批可能要等XX天。活动价随时可能结束, 要拍的宝宝抓紧时间下单。\n\n【关键点】制造稀缺+紧迫感, 但注意合规(不要虚假库存)。',
     '30秒-1分钟'),
    ('Step 7', '转品/下播总结',
     '"拍了Xiaomi Buds 6的宝宝, 可以再看看我们的开放式耳机, 适合运动跑步、长时间佩戴不闷耳朵。运动+办公双场景, 音质续航同样出色。\n\n还没关注的宝宝点个关注, 明天X点我们准时开播, 来就有福利。"\n\n【关键点】关联推荐提升客单价, 引导关注提高复购。',
     '1分钟'),
]

for step, module, script, timing in sop_items:
    sc(ws4, row, 1, step, font=bold_font)
    sc(ws4, row, 2, module, font=bold_font)
    sc(ws4, row, 3, script, font=Font(name='Arial', size=9), align=left_align)
    sc(ws4, row, 4, timing, font=bold_font)
    ws4.row_dimensions[row].height = 130
    row += 1

widths4 = [10, 22, 90, 14]
for i, w in enumerate(widths4):
    ws4.column_dimensions[get_column_letter(i+1)].width = w

# Key techniques section
row += 1
row = add_section_header(ws4, row, '话术关键技巧 (6大核心)', MAX_C4)
techniques = [
    ('互动确认', '通过"要的飘1""已拍的飘已拍"等互动确认用户意向再推进话术, 避免自说自话'),
    ('适配度先行', '在讲任何产品卖点之前, 先打消"我的手机能不能用"的顾虑, 这是最高频的流失原因'),
    ('场景化描述', '"坐地铁时""打游戏时""跑步时""开会时"——把功能融入用户真实生活场景, 增强代入感'),
    ('限时紧迫感', '"只剩X个""活动即将结束""这批卖完等X天"——合法合理的稀缺制造'),
    ('精准推品', '按预算分层推荐: 不差钱推旗舰, 性价比推中端, 节约推入门。推荐逻辑要有说服力'),
    ('售后锁单', '官方+7天无理由+1年质保+顺丰=四重保障降低决策门槛, 这是官方直播间最大优势'),
]
for title, desc in techniques:
    ws4.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_C4)
    sc(ws4, row, 1, f'{title}: {desc}', font=Font(name='Arial', size=10), align=left_align)
    ws4.row_dimensions[row].height = 22
    row += 1

# Common objections
row += 1
row = add_section_header(ws4, row, '耳机直播间 常见异议处理', MAX_C4)

objections = [
    ('"苹果手机能用吗?"', '可以用! 只要是智能手机有蓝牙就能连。苹果/华为/OPPO/vivo一视同仁。小米怎么连, 苹果就怎么连, 放心拍。'),
    ('"续航多久? 够用吗?"', 'Buds 6单耳10小时(关降噪)/6小时(开降噪), 充电盒再续30小时。一周充一次电, 短期出差不用带充电线。'),
    ('"声音会不会有延迟?"', 'LDHC 5.0+蓝牙5.4, 延迟低至55ms, 打游戏/看视频基本感知不到延迟。支持低延迟游戏模式。'),
    ('"半入耳降噪效果好吗?"', '半入耳确实不如入耳式物理隔音好, 但Buds 6的AI自适应降噪在地铁/公交/办公室场景效果很好, 日常使用完全够。追求极致降噪可以看我们的入耳式。'),
    ('"跟XX品牌比哪个好?"', '不贬低竞品。讲自己优势: Hi-Res认证/独立空间音频/40小时续航/双设备切换, 这个价位段Buds 6的配置是最全的。您可以去对比。'),
    ('"为什么比其他平台便宜?"', '我们是小米官方直播间, 厂家直销没有中间商赚差价。正品保障+官方发票+全国联保, 比其他渠道更放心。'),
    ('"售后怎么保障?"', '7天无理由(激活也能退)+1年质保全国联保+顺丰京东配送。任何问题去就近小米门店免费修。'),
    ('"送人合适吗? 包装怎么样?"', '小米官方包装精美, 送礼体面。耳机是数码礼品里的实用选择, 性价比高还不容易踩雷。'),
]
for q, a in objections:
    rf = alt_fill if row % 2 == 0 else white_fill
    # Write to separate columns without merging
    sc(ws4, row, 1, q, font=bold_font, align=left_align, fill=rf)
    sc(ws4, row, 2, '', fill=rf)
    ws4.merge_cells(start_row=row, start_column=3, end_row=row, end_column=MAX_C4)
    sc(ws4, row, 3, a, font=Font(name='Arial', size=9), align=left_align, fill=rf)
    for cc in range(3, MAX_C4+1):
        ws4.cell(row=row, column=cc).fill = rf
        ws4.cell(row=row, column=cc).border = thin_border
    ws4.row_dimensions[row].height = 30
    row += 1

print('Sheet 4 done')

# ============================================================
# SHEET 5: 节奏分析
# ============================================================
ws5 = wb.create_sheet('节奏与排班分析')
MAX_C5 = 6

row = 1
row = add_title(ws5, row, '直播节奏分析与优化建议', MAX_C5)

# Current rhythm
row += 1
row = add_section_header(ws5, row, '当前节奏诊断', MAX_C5)

rhythm_issues = [
    ('问题1: 排班不规律', '7.1-7.12每天开播但产出不稳定(3K-24K波动); 7.13-7.15可能断播; 7.16-7.18恢复。不规律的直播节奏影响粉丝积累和平台推流权重。'),
    ('问题2: 主播结构失衡', '2个S级+2个A/B级+3个C级新人。S级仅播4-5天, A/B级播12天。流量高峰期(如7.17爆品日)S级主播上线后效果立竿见影, 说明平时主播能力不足导致流量浪费。'),
    ('问题3: 时段选择不明', '耳机直播间的最佳开播时段不明确。从数据看, 7.17/7.18爆发可能与平台流量高峰+爆品节点相关, 建议分析具体时段数据。'),
    ('问题4: 没有固定栏目', '缺少固定的节目节奏: 如新品首发/对比测评/粉丝福利日等。内容单一导致老粉流失快。'),
]
for title, desc in rhythm_issues:
    ws5.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_C5)
    sc(ws5, row, 1, f'{title}: {desc}', font=Font(name='Arial', size=10), align=left_align)
    ws5.row_dimensions[row].height = 28
    row += 1

# Optimized schedule
row += 1
row = add_section_header(ws5, row, '优化后的直播排班建议 (7.21-7.27)', MAX_C5)

sched_headers = ['时段', '时间', '主播安排', '内容规划', '目标GSV', '备注']
for i, h in enumerate(sched_headers):
    sc(ws5, row, i+1, h, font=header_font, fill=header_fill2)
ws5.row_dimensions[row].height = 28
row += 1

schedule = [
    ('早场(测试)', '08:00-11:00', '李晓洋 / 王瑞(轮换)', '产品讲解+答疑, 侧重REDMI Buds系列, 拉新为主', '3,000-5,000', '低流量时段, 适合腰部主播练手+承接自然流量'),
    ('午场(主推)', '12:00-15:00', '王瑞 / 范金蝶(轮换)', 'Xiaomi Buds 6主推+开放式耳机, 配合午休流量高峰', '5,000-10,000', '午休时段流量较好, 安排有经验的主播转化'),
    ('下午场(蓄水)', '15:00-18:00', '李晓洋(固定)+新人旁听', '对比测评+互动问答, 新人逐步参与讲解', '3,000-5,000', '新人培养时段, 以老带新'),
    ('晚场(爆发)', '19:00-23:00', '徐梦洁(优先)/刘子源(轮换)', '爆品主推+限时活动, Xiaomi Buds 6+高端品, 全力冲GSV', '15,000-50,000', '流量黄金时段, 必须排S级主播, 配合投流和活动'),
    ('深夜场(增量)', '23:00-01:00', '刘子源 / 范金蝶(轮换)', '深夜福利+性价比产品, REDMI Buds 8系列走量', '5,000-10,000', '深夜竞争少, 适合走量型主播'),
]
for s in schedule:
    for i, v in enumerate(s):
        if i == 3:
            sc(ws5, row, i+1, v, font=Font(name='Arial', size=9), align=left_align)
        elif i == 4:
            sc(ws5, row, i+1, v, font=highlight_font)
        else:
            sc(ws5, row, i+1, v, font=normal_font)
    ws5.row_dimensions[row].height = 35
    row += 1

widths5 = [14, 14, 22, 40, 16, 40]
for i, w in enumerate(widths5):
    ws5.column_dimensions[get_column_letter(i+1)].width = w

# Rhythm SOP
row += 1
row = add_section_header(ws5, row, '单场直播节奏SOP (以3小时为例)', MAX_C5)

rhythm_sop = [
    ('0-15分钟', '暖场 + 福利预告', '欢迎新老粉, 告知今天主推品和福利, 引导点赞/关注/加粉丝团。快速回答弹幕高频问题。'),
    ('15-45分钟', '主打品深度讲解', '选择1-2个主推SKU深度塑品(用Step 1-3话术)。现场演示连接/音质/降噪效果。'),
    ('45-60分钟', '互动 + 促单', '回答弹幕问题, 引导已拍扣"已拍", 用库存+优惠逼单。'),
    ('60-90分钟', '次推品讲解', '讲解第二个主推品或配件。关联推荐: 拍耳机的推荐充电配件/耳机套。'),
    ('90-105分钟', '互动游戏 + 粉丝福利', '抽奖/秒杀/粉丝专属价, 提升互动率和停留时长。'),
    ('105-135分钟', '再次主推品逼单', '回到主打品, 用"最后X单""活动即将结束"再次逼单。强调售后保障。'),
    ('135-165分钟', '全品回顾 + 答疑', '快速回顾今天所有品, 集中答疑, 处理售后问题。收集粉丝反馈。'),
    ('165-180分钟', '下播预告 + 关注引导', '预告明天直播内容和福利, 引导关注/开播提醒。感谢打赏和下单粉丝。'),
]
for time_slot, module, detail in rhythm_sop:
    sc(ws5, row, 1, time_slot, font=bold_font)
    sc(ws5, row, 2, module, font=bold_font)
    sc(ws5, row, 3, detail, font=Font(name='Arial', size=9), align=left_align)
    for c in range(4, MAX_C5+1):
        sc(ws5, row, c, '', fill=white_fill)
    ws5.row_dimensions[row].height = 28
    row += 1

# Merge for the description columns
ws5.merge_cells(start_row=row-8, start_column=3, end_row=row-1, end_column=MAX_C5)

print('Sheet 5 done')

# ============================================================
# SHEET 6: 下周规划
# ============================================================
ws6 = wb.create_sheet('下周规划(7.21-7.27)')
MAX_C6 = 5

row = 1
row = add_title(ws6, row, '下周工作规划 (7月21日-7月27日)', MAX_C6)
row = add_subtitle(ws6, row, '基于7月1-18日数据复盘制定 | 目标: 全周GSV突破20万, 缩小与良米差距', MAX_C6)

# Goals
row += 1
row = add_section_header(ws6, row, '一、核心目标', MAX_C6)
goals = [
    ('GSV目标', '全周GSV 200,000+ (日均28,571, 较上周均值增长30%)'),
    ('订单目标', '全周订单 400+ (日均57单, 客单价维持450+)'),
    ('主播目标', 'S级主播(徐梦洁/刘子源)各播5天以上; A/B级主播日均3,500+; 新人完成3天旁听+2天试播'),
    ('排班目标', '每天保证至少2场直播(白天+晚场), 总直播时长50小时+/周'),
    ('产品目标', 'Xiaomi Buds 6占比提升至40%+; 客单价维持450+; 客单价不降的同时冲量'),
]
for g, detail in goals:
    ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_C6)
    sc(ws6, row, 1, f'{g}: {detail}', font=Font(name='Arial', size=10), align=left_align)
    ws6.row_dimensions[row].height = 24
    row += 1

# Weekly schedule
row += 1
row = add_section_header(ws6, row, '二、排班计划', MAX_C6)

day_headers = ['日期', '星期', '早场(8-11)', '午场(12-15)', '晚场(19-23)', '深夜场(23-01)']
for i, h in enumerate(day_headers):
    sc(ws6, row, i+1, h, font=header_font, fill=header_fill2)
ws6.row_dimensions[row].height = 28
row += 1

days_schedule = [
    ('7.21', '一', '李晓洋', '王瑞', '徐梦洁', '休息'),
    ('7.22', '二', '李晓洋+于飞(旁听)', '范金蝶', '刘子源', '王瑞'),
    ('7.23', '三', '王瑞', '李晓洋', '徐梦洁', '休息'),
    ('7.24', '四', '李晓洋+郑涵悦(旁听)', '范金蝶', '刘子源', '范金蝶'),
    ('7.25', '五', '王瑞', '李晓洋', '徐梦洁', '休息'),
    ('7.26', '六', '李晓洋+王洋帆(旁听)', '范金蝶', '刘子源+王瑞', '王瑞'),
    ('7.27', '日', '王瑞', '徐梦洁', '徐梦洁+刘子源', '休息'),
]
for d in days_schedule:
    for i, v in enumerate(d):
        if i >= 2 and '徐梦洁' in v or '刘子源' in v:
            sc(ws6, row, i+1, v, font=Font(name='Arial', size=10, bold=True, color='ff6900'), fill=gold_fill)
        elif i >= 2 and '旁听' in v:
            sc(ws6, row, i+1, v, font=Font(name='Arial', size=9, color='3b82f6'), fill=blue_fill)
        elif i >= 2 and v == '休息':
            sc(ws6, row, i+1, v, font=Font(name='Arial', size=9, color='999999'), fill=light_gray_fill)
        else:
            sc(ws6, row, i+1, v, font=normal_font)
    ws6.row_dimensions[row].height = 28
    row += 1

widths6 = [10, 8, 20, 16, 22, 22]
for i, w in enumerate(widths6):
    ws6.column_dimensions[get_column_letter(i+1)].width = w

# Key actions
row += 1
row = add_section_header(ws6, row, '三、重点行动项', MAX_C6)

actions = [
    ('1. 话术培训 (周一)', '组织全体主播话术培训, 重点:\n- 学习徐梦洁/刘子源7.17/7.18高转化场的录屏\n- 演练7步SOP话术流程\n- 重点突破: 逼单话术、异议处理、客单价拉升技巧\n- 每人模拟直播5分钟, 互相点评'),
    ('2. 爆品策略 (全周)', 'Xiaomi Buds 6作为主打品, 所有场次必须讲解至少1次\n- 准备对比道具(竞品耳机/音质对比视频)\n- 录制产品演示短视频发抖音/视频号引流\n- 设置主推品专属优惠码"Buds6专属"'),
    ('3. 新人培养 (全周)', '于飞/郑涵悦/王洋帆:\n- 前3天: 旁听+记笔记+模拟演练\n- 第4-5天: 在师父陪同下试播30分钟\n- 第6-7天: 独立播1小时, 师父在旁指导\n- 考核标准: 独立场次GSV>1,000即为通过'),
    ('4. 数据监控 (每日)', '每日播后复盘:\n- 记录每场GSV/订单/客单价/新增粉丝/观看人数\n- 对比目标差距, 当天调整次日策略\n- 每周五汇总周报, 分析趋势和问题'),
    ('5. 竞品监控 (持续)', '重点关注良米(小米耳机官方直播间):\n- 记录他们的排班/话术/活动/产品组合\n- 分析他们7.10-7.12高转化(日均40K+)的原因\n- 吸收可复用的策略'),
    ('6. 流量获取 (全周)', '增加直播间流量:\n- 每天发布1-2条短视频(产品测评/使用场景)\n- 晚场配合投流(预算建议200-500/天)\n- 引导粉丝转发直播间到微信群/朋友圈\n- 与小米其他官方直播间互推'),
]
for title, detail in actions:
    ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_C6)
    sc(ws6, row, 1, f'{title}', font=bold_font, align=left_align)
    ws6.row_dimensions[row].height = 20
    row += 1
    ws6.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_C6)
    sc(ws6, row, 1, detail, font=Font(name='Arial', size=9), align=left_align)
    ws6.row_dimensions[row].height = 65
    row += 1

# KPIs
row += 1
row = add_section_header(ws6, row, '四、下周KPI考核', MAX_C6)

kpi_headers = ['指标', '目标值', '考核周期', '责任人', '奖惩说明']
for i, h in enumerate(kpi_headers):
    sc(ws6, row, i+1, h, font=header_font, fill=header_fill)
ws6.row_dimensions[row].height = 28
row += 1

kpis = [
    ('全直播间周GSV', '200,000+', '7.21-7.27', '全体主播', '达标: 团队奖金池; 超标120%: 额外奖励'),
    ('S级主播场均GSV', '15,000+', '每场', '徐梦洁/刘子源', '低于10K复盘分析; 连续3场低于10K调整排班'),
    ('A/B级主播日均GSV', '3,500+', '每日', '王瑞/李晓洋/范金蝶', '达标率>80%合格; <60%参加额外培训'),
    ('新人首周GSV', '1,000+/人', '试播日', '于飞/郑涵悦/王洋帆', '达标后可独立排班; 未达标延长带教期1周'),
    ('客单价', '450+', '每周均值', '全体主播', '低于400需检查是否过度推低端品'),
    ('新增粉丝', '200+/周', '7.21-7.27', '全体主播', '每场开播引导关注; 短视频引流到直播间'),
]
for k in kpis:
    for i, v in enumerate(k):
        sc(ws6, row, i+1, v, font=normal_font, align=left_align if i in (3,4) else center_align)
    ws6.row_dimensions[row].height = 26
    row += 1

print('Sheet 6 done')

# ============================================================
# SHEET 7: 我司 vs 良米 对比
# ============================================================
ws7 = wb.create_sheet('我司vs良米对比')
MAX_C7 = 10

row = 1
row = add_title(ws7, row, '我司(小米官方耳机直播间) vs 良米(小米耳机官方直播间) 7月对比', MAX_C7)

# Summary KPIs
row += 1
row = add_section_header(ws7, row, '核心指标对比', MAX_C7)

comp_headers = ['指标', '我司', '良米', '差距', '解读']
for i, h in enumerate(comp_headers):
    sc(ws7, row, i+1, h, font=header_font, fill=header_fill2)
ws7.row_dimensions[row].height = 28
row += 1

our_total_orders = 759
our_total_rev = 360134
lm_total_orders = 4617
lm_total_rev = 1756223

comp_data = [
    ('总订单', f'{our_total_orders}单', f'{lm_total_orders}单', f'差{lm_total_orders-our_total_orders}单 ({our_total_orders/lm_total_orders*100:.0f}%)', '良米订单量是我司的6.1倍, 体量差距悬殊'),
    ('总GSV', f'{our_total_rev:,}', f'{lm_total_rev:,}', f'差{lm_total_rev-our_total_rev:,} ({our_total_rev/lm_total_rev*100:.0f}%)', '良米GSV是我司的4.9倍'),
    ('客单价', f'{our_total_rev/our_total_orders:.0f}', f'{lm_total_rev/lm_total_orders:.0f}', f'高{our_total_rev/our_total_orders - lm_total_rev/lm_total_orders:.0f}', '我司客单价高94元(474 vs 380), 产品结构更高端'),
    ('日均订单', f'{our_total_orders/19:.0f}单', f'{lm_total_orders/19:.0f}单', f'差{lm_total_orders/19 - our_total_orders/19:.0f}单', '良米日均243单 vs 我司40单'),
    ('日均GSV', f'{our_total_rev/19:,.0f}', f'{lm_total_rev/19:,.0f}', f'差{lm_total_rev/19 - our_total_rev/19:,.0f}', '良米日均92K vs 我司19K'),
    ('市场份额', '17.0%', '83.0%', '-66%', '我司仅占两个耳机直播间的17%, 提升空间巨大'),
    ('主播数(估)', '9人', '约12-15人', '差3-6人', '良米主播团队更大, 排班更密集'),
    ('日均直播时长(估)', '约3-4小时', '约6-8小时', '差3-4小时', '良米直播时长更长, 覆盖更多时段'),
]
for item in comp_data:
    for i, v in enumerate(item):
        if i == 1:
            sc(ws7, row, i+1, v, font=Font(name='Arial', size=10, bold=True, color='1E90FF'))
        elif i == 2:
            sc(ws7, row, i+1, v, font=Font(name='Arial', size=10, bold=True, color='FF6B35'))
        elif i == 4:
            sc(ws7, row, i+1, v, font=small_font, align=left_align)
        else:
            sc(ws7, row, i+1, v, font=normal_font)
    ws7.row_dimensions[row].height = 26
    row += 1

# Daily comparison table
row += 1
row = add_section_header(ws7, row, '逐日对比 (7.1-7.19)', MAX_C7)

daily_comp_headers = ['日期', '我司订单', '我司GSV', '我司客单价', '良米订单', '良米GSV', '良米客单价', '我司GSV占比', '趋势']
# We need 9 columns for this but MAX_C7 is 10

# Build daily comparison from room_comp data
row_header = row
for date_idx, d in enumerate(july_data):
    date_str = d['date'][5:]
    rooms = d.get('rooms', {})
    our = rooms.get('小米官方耳机直播间', {})
    lm = rooms.get('小米耳机官方直播间', {})

    our_orders = our.get('orders', 0)
    our_rev = our.get('revenue', 0)
    lm_orders = lm.get('orders', 0)
    lm_rev = lm.get('revenue', 0)

    our_asp = our_rev / our_orders if our_orders > 0 else 0
    lm_asp = lm_rev / lm_orders if lm_orders > 0 else 0
    share = our_rev / (our_rev + lm_rev) * 100 if (our_rev + lm_rev) > 0 else 0

    if share >= 30:
        trend_icon = '📈 领先'
        trend_f = 'f0faf3'
    elif share >= 15:
        trend_icon = '📊 追赶中'
        trend_f = 'fff7ed'
    elif share > 0:
        trend_icon = '⚠️ 落后'
        trend_f = 'fef2f2'
    else:
        trend_icon = '❌ 未开播'
        trend_f = 'f5f5f5'

    rf = alt_fill if date_idx % 2 == 0 else white_fill

    vals = [date_str, our_orders, our_rev, our_asp, lm_orders, lm_rev, lm_asp, f'{share:.1f}%', trend_icon]
    for i, v in enumerate(vals):
        if i in (2, 5):
            sc(ws7, row, i+1, v, font=bold_font, nf='#,##0', fill=rf)
        elif i in (3, 6):
            sc(ws7, row, i+1, v, font=bold_font, nf='#,##0', fill=rf)
        elif i == 7:
            sc(ws7, row, i+1, v, font=Font(name='Arial', size=10, bold=True, color='ff6900' if share>=20 else '333333'), fill=rf)
        elif i == 8:
            sc(ws7, row, i+1, v, font=small_font, fill=PatternFill('solid', fgColor=trend_f))
        else:
            sc(ws7, row, i+1, v, fill=rf)
    ws7.row_dimensions[row].height = 22
    row += 1

# Add header row
row_h = row_header
daily_headers = ['日期', '我司订单', '我司GSV', '我司客单价', '良米订单', '良米GSV', '良米客单价', '我司GSV占比', '趋势']
ws7.insert_rows(row_h)
for i, h in enumerate(daily_headers):
    sc(ws7, row_h, i+1, h, font=header_font, fill=header_fill)
ws7.row_dimensions[row_h].height = 28

widths7 = [10, 10, 12, 10, 10, 12, 10, 12, 12]
for i, w in enumerate(widths7):
    ws7.column_dimensions[get_column_letter(i+1)].width = w
# Add extra column
ws7.column_dimensions[get_column_letter(10)].width = 40

# Strategic recommendations
row += 1
row = add_section_header(ws7, row, '追赶良米的策略建议', MAX_C7)

strategies = [
    '1. 差异化定位: 良米走量(客单价380), 我司走质(客单价474)。坚持高端产品线(Xiaomi Buds 6/开放式耳机)为主, 不做低价竞争。',
    '2. 增加直播时长: 从当前日均3-4小时扩展到6-8小时, 覆盖更多时段。至少保证午场+晚场+深夜场3个时段。',
    '3. 扩大主播团队: 在现有9人基础上, 重点培养3个新人尽快独立上岗, 考虑再招募1-2名有经验的主播。',
    '4. 爆品日策略: 学习7.17模式(徐梦洁+刘子源双S级+Buds 6爆品), 每周至少安排1次"超级爆品日"。',
    '5. 投流配合: 良米可能已在投流获客。建议7.21起每天晚场配合投流200-500元, 测试ROI后逐步加量。',
    '6. 视频引流: 每天发布产品测评/使用教程短视频, 从抖音/视频号引流到直播间。良米大概率已在做。',
]
for s in strategies:
    ws7.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_C7)
    sc(ws7, row, 1, s, font=Font(name='Arial', size=10), align=left_align)
    ws7.row_dimensions[row].height = 24
    row += 1

print('Sheet 7 done')

# ============================================================
# SHEET 8: 产品销售分析
# ============================================================
ws8 = wb.create_sheet('产品销售分析')
MAX_C8 = 8

row = 1
row = add_title(ws8, row, '7月耳机产品销量分析 (我司直播间)', MAX_C8)
row = add_subtitle(ws8, row, '产品销售排名 | 客单价分析 | 产品结构优化建议', MAX_C8)

# Aggregate product sales from our room
from collections import defaultdict
our_products = defaultdict(lambda: {'orders': 0, 'revenue': 0})
lm_products = defaultdict(lambda: {'orders': 0, 'revenue': 0})

for d in july_data:
    rooms = d.get('rooms', {})
    our = rooms.get('小米官方耳机直播间', {})
    our_prods = our.get('products', {})
    for pn, pd in our_prods.items():
        our_products[pn]['orders'] += pd.get('orders', 0)
        our_products[pn]['revenue'] += pd.get('revenue', 0)

    lm = rooms.get('小米耳机官方直播间', {})
    lm_prods = lm.get('products', {})
    for pn, pd in lm_prods.items():
        lm_products[pn]['orders'] += pd.get('orders', 0)
        lm_products[pn]['revenue'] += pd.get('revenue', 0)

row += 1
row = add_section_header(ws8, row, '我司产品销售排行 (7.1-7.19)', MAX_C8)

prod_headers = ['排名', '产品名称', '销量(单)', '销售额', '客单价', '占我司耳机GSV比', '产品定位', '建议']
for i, h in enumerate(prod_headers):
    sc(ws8, row, i+1, h, font=header_font, fill=header_fill2)
ws8.row_dimensions[row].height = 28
row += 1

sorted_our = sorted(our_products.items(), key=lambda x: x[1]['revenue'], reverse=True)
total_our_prod_rev = sum(v['revenue'] for v in our_products.values())

for rank, (pn, pd) in enumerate(sorted_our):
    rev = pd['revenue']
    orders = pd['orders']
    asp = rev / orders if orders > 0 else 0
    pct = rev / total_our_prod_rev * 100 if total_our_prod_rev > 0 else 0

    if 'Buds 6' in pn and 'REDMI' not in pn:
        pos = '旗舰主推品'
        suggestion = '全力主推, 所有场次必讲'
        sug_fill = gold_fill
    elif '开放式' in pn:
        pos = '高端差异化'
        suggestion = '运动/办公场景重点推, 提升客单价'
        sug_fill = green_fill
    elif '手环' in pn:
        pos = '跨品类搭配'
        suggestion = '耳机用户交叉推荐, 提升连带率'
        sug_fill = blue_fill
    elif 'Pro' in pn:
        pos = '中端走量'
        suggestion = '性价比用户首选, 配合主推品做对比'
        sug_fill = white_fill
    elif '青春版' in pn or '活力版' in pn:
        pos = '入门引流'
        suggestion = '引流款, 不重点推但要有库存'
        sug_fill = light_gray_fill
    else:
        pos = '补充品类'
        suggestion = '按需求推荐'
        sug_fill = white_fill

    rf = alt_fill if rank % 2 == 0 else white_fill
    vals = [rank+1, pn, orders, rev, asp, f'{pct:.1f}%', pos, suggestion]
    for i, v in enumerate(vals):
        if i == 3:
            sc(ws8, row, i+1, v, font=highlight_font, nf='#,##0', fill=rf)
        elif i == 4:
            sc(ws8, row, i+1, v, nf='#,##0', fill=rf)
        elif i == 6:
            sc(ws8, row, i+1, v, font=bold_font, fill=rf)
        elif i == 7:
            sc(ws8, row, i+1, v, font=small_font, align=left_align, fill=sug_fill)
        else:
            sc(ws8, row, i+1, v, fill=rf)
    ws8.row_dimensions[row].height = 24
    row += 1

# Product strategy
row += 1
row = add_section_header(ws8, row, '产品结构优化策略', MAX_C8)

prod_strategies = [
    ('主推品 (占GSV 60%+)', 'Xiaomi Buds 6 + Xiaomi 开放式耳机', '全时段全主播必讲, 准备对比道具, 设置专属优惠码, 录制演示视频引流'),
    ('利润品 (占GSV 20-30%)', 'REDMI Buds 8 Pro + REDMI Buds 8', '作为"性价比之选"在主推品之后讲解, 拉高转化率, 走量赚利润'),
    ('引流品 (占GSV 5-10%)', 'REDMI Buds 8青春版/活力版', '低价引流, 吸引价格敏感用户进入直播间, 再向上转化到Pro或Buds 6'),
    ('搭配品 (提升客单价)', '充电配件/耳机套/耳机挂绳', '每个拍耳机的用户推荐搭配, 成本低利润高, 有效提升客单价'),
    ('跨品类 (提升连带)', '小米手环10/手表', '耳机+手环/手表场景联动推荐(运动时同时用), 提升单用户价值'),
]
for title, products, strategy in prod_strategies:
    ws8.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_C8)
    sc(ws8, row, 1, f'{title}: {products} -> {strategy}', font=Font(name='Arial', size=10), align=left_align)
    ws8.row_dimensions[row].height = 24
    row += 1

widths8 = [8, 30, 10, 14, 10, 14, 16, 50]
for i, w in enumerate(widths8):
    ws8.column_dimensions[get_column_letter(i+1)].width = w

# 良米 product comparison
row += 1
row = add_section_header(ws8, row, '良米热销产品参考 (7.1-7.19, 仅前5)', MAX_C8)

sorted_lm = sorted(lm_products.items(), key=lambda x: x[1]['revenue'], reverse=True)
for rank, (pn, pd) in enumerate(sorted_lm[:5]):
    rev = pd['revenue']
    orders = pd['orders']
    asp = rev / orders if orders > 0 else 0
    sc(ws8, row, 1, rank+1)
    sc(ws8, row, 2, pn)
    sc(ws8, row, 3, orders, font=Font(name='Arial', size=10, bold=True, color='FF6B35'))
    sc(ws8, row, 4, rev, font=Font(name='Arial', size=10, bold=True, color='FF6B35'), nf='#,##0')
    sc(ws8, row, 5, asp, nf='#,##0')
    for c in range(6, MAX_C8+1):
        sc(ws8, row, c, '')
    ws8.row_dimensions[row].height = 22
    row += 1

row += 1
ws8.merge_cells(start_row=row, start_column=1, end_row=row, end_column=MAX_C8)
sc(ws8, row, 1, '注意: 良米可能以REDMI Buds 8青春版/活力版为主要走量品(客单价低), 我司应坚持高端差异化路线, 不盲目跟进低价竞争。', font=Font(name='Arial', size=9, color='999999', italic=True), align=left_align)

print('Sheet 8 done')

# ============================================================
# FINAL SAVE
# ============================================================
output_path = r'C:\Users\Administrator\Desktop\小米耳机直播间_7月主播复盘报告.xlsx'
wb.save(output_path)
print(f'\n报告已生成: {output_path}')
print(f'共8个Sheet: {", ".join(wb.sheetnames)}')
