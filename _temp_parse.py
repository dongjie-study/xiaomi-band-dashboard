import openpyxl, json, re

wb = openpyxl.load_workbook(r'C:\Users\Administrator\Desktop\8.10业绩.xlsx')
ws = wb.active

# Room name → roomId mapping (from WORKFLOW.md)
NAME_TO_ROOMID = {
    '小米数码旗舰店': 'room_xiaomi_digital',
    '小米官方手环直播间': 'room_xiaomi_band',
    '小米官方手表': 'room_xiaomi_watch',
    '小米官旗手表直播间': 'room_xiaomi_watch_flagship',
    '小米官方耳机直播间': 'room_xiaomi_earphone',
    '小米AI眼镜直播间': 'room_xiaomi_glasses',
    '小米智能设备旗舰店直播间': 'room_xiaomi_smart_device',
    '小米手环官旗直播间': 'room_xiaomi_band_flagship',
}

SHIFT_MAP = {'A1': 'A', 'A2': 'A', 'B1': 'B', 'B2': 'B', 'C1': 'C', 'C2': 'C', 'D1': 'D', 'D2': 'D', 'E': 'E'}

# Find room header rows: col A has name, col B is None, col C is None
room_rows = []
for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
    a = row[0].value
    b = row[1].value
    c = row[2].value
    if a and str(a).strip() != '主播' and b is None and c is None:
        room_rows.append((row[0].row, str(a).strip()))

print(f"Found {len(room_rows)} rooms:")
records = []

for i, (start_row, room_name) in enumerate(room_rows):
    # Match room name to roomId
    room_id = None
    for key, rid in NAME_TO_ROOMID.items():
        if key in room_name or room_name in key:
            room_id = rid
            break

    if not room_id:
        # Try fuzzy match
        print(f"  WARNING: Unknown room '{room_name}' at row {start_row}, trying to match...")
        # Fallback: use position-based mapping
        pos_map = [
            'room_xiaomi_digital', 'room_xiaomi_band', 'room_xiaomi_watch',
            'room_xiaomi_watch_flagship', 'room_xiaomi_earphone', 'room_xiaomi_glasses',
            'room_xiaomi_smart_device', 'room_xiaomi_band_flagship'
        ]
        if i < len(pos_map):
            room_id = pos_map[i]
            print(f"  → fallback to {room_id}")
        else:
            continue

    print(f"  {room_name} → {room_id}")

    end_row = room_rows[i + 1][0] - 1 if i + 1 < len(room_rows) else ws.max_row

    # Scan anchor rows within this room
    for r in range(start_row + 1, end_row + 1):
        a_val = ws.cell(row=r, column=1).value      # col A: anchor name
        shift_val = ws.cell(row=r, column=3).value    # col C: shift (A1, B1, E, etc.)
        total_val = ws.cell(row=r, column=6).value    # col F: total GSV

        if not a_val or str(a_val).strip() == '':
            continue
        if not shift_val or str(shift_val).strip() not in SHIFT_MAP:
            continue

        anchor = str(a_val).strip()
        shift = SHIFT_MAP[str(shift_val).strip()]

        # Compute total sales
        if isinstance(total_val, str) and total_val.startswith('='):
            m = re.match(r'=SUM\(E(\d+):E(\d+)\)', total_val)
            if m:
                e_s, e_e = int(m.group(1)), int(m.group(2))
                total = sum(float(ws.cell(row=er, column=5).value or 0) for er in range(e_s, e_e + 1))
            else:
                total = 0
        elif isinstance(total_val, (int, float)):
            total = float(total_val)
        else:
            total = 0

        records.append({
            'roomId': room_id,
            'shift': shift,
            'anchor': anchor,
            'sales': round(total, 2)
        })

print(f"\n{len(records)} records extracted:")
for r in records:
    print(f"  {r['roomId']} | {r['shift']} | {r['anchor']} | {r['sales']}")

with open(r'C:\Users\Administrator\Desktop\temp_records.json', 'w', encoding='utf-8') as f:
    json.dump(records, f, ensure_ascii=False, indent=2)
print("\nDone.")
